"""
employees_to_db.py  —  Base uploader for the 'employees' collection.
Clears the collection and inserts all rows from employees_data.xlsx.

Usage:
    python Uploader_Scripts/Base_Uploader_Scripts/employees_to_db.py
"""

import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

import openpyxl
from pymongo import MongoClient
from datetime import datetime

MONGO_URI    = "mongodb://localhost:27017/"
DB_NAME      = "test_db"
COLLECTION   = "employees"
EXCEL_FILE   = os.path.join(os.path.dirname(__file__), "..", "..", "input_folder", "employees_data.xlsx")
SHEET_NAME   = "Employees"
HEADER_ROW   = 1


def sanitize_key(k):
    return k.replace(".", "_").replace("$", "_") if k else "col"


def read_sheet(path, sheet, header_row):
    ws = openpyxl.load_workbook(path)[sheet]
    headers = [sanitize_key(str(c.value)) if c.value else f"col{i}"
               for i, c in enumerate(ws[header_row])]
    headers.append("date")
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row) + [datetime.now()]
        rows.append(dict(zip(headers, vals)))
    return rows


def main():
    data = read_sheet(EXCEL_FILE, SHEET_NAME, HEADER_ROW)
    col  = MongoClient(MONGO_URI)[DB_NAME][COLLECTION]
    col.delete_many({})
    if data:
        col.insert_many(data)
    print(f"[{COLLECTION}]  Inserted {len(data)} records (collection cleared first)")


if __name__ == "__main__":
    main()
