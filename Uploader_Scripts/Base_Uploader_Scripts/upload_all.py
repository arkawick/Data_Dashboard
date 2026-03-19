"""
upload_all.py  —  CLI batch uploader (no GUI)
Reads all 5 Excel files from input_folder/ and inserts them into MongoDB.
Requirements collection uses sync (tracks added/updated/deleted).

Usage:
    python Uploader_Scripts/Base_Uploader_Scripts/upload_all.py
"""

import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

import re
import openpyxl
from pymongo import MongoClient
from datetime import datetime
from config import FILE_PATTERNS

MONGO_URI    = "mongodb://localhost:27017/"
DB_NAME      = "test_db"
SCRIPT_DIR   = os.path.join(os.path.dirname(__file__), "..", "..")
INPUT_FOLDER = os.path.join(SCRIPT_DIR, "input_folder")


def get_col(name):
    return MongoClient(MONGO_URI)[DB_NAME][name]


def sanitize_key(k):
    return k.replace(".", "_").replace("$", "_") if k else "col"


def sanitize_headers(hdrs):
    seen, out = {}, []
    for h in hdrs:
        s = sanitize_key(str(h)) if h else "col"
        s = f"{s}_{seen[s]}" if s in seen and (seen.update({s: seen[s]+1}) or True) else s
        seen.setdefault(s, 0)
        out.append(s)
    return out


def read_sheet(path, sheet, header_row=1):
    ws = openpyxl.load_workbook(path)[sheet]
    hdrs = [c.value for c in ws[header_row]]
    hdrs_s = sanitize_headers(hdrs)
    hdrs_s.append("date")
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row) + [datetime.now()]
        rows.append(dict(zip(hdrs_s, vals)))
    return rows


def insert_fresh(data, col):
    col.delete_many({})
    if data:
        col.insert_many(data)
    print(f"  [{col.name}]  inserted {len(data)} records (collection cleared)")


def sync_col(data, col, uid_field):
    existing = {d[uid_field]: d for d in col.find({}, {uid_field: 1, "_id": 0}) if d.get(uid_field)}
    seen, added, updated = set(), 0, 0
    for rec in data:
        uid = rec.get(uid_field)
        if not uid:
            continue
        if uid in existing:
            col.update_one({uid_field: uid}, {"$set": {**rec, "db_status": "updated"}})
            updated += 1
        else:
            col.insert_one({**rec, "db_status": "added"})
            added += 1
        seen.add(uid)
    deleted = col.update_many({uid_field: {"$nin": list(seen)}},
                               {"$set": {"db_status": "deleted"}}).modified_count
    print(f"  [{col.name}]  synced: {added} added, {updated} updated, {deleted} deleted")


def main():
    files = sorted(f for f in os.listdir(INPUT_FOLDER) if f.endswith(".xlsx"))
    if not files:
        print("No Excel files in input_folder/. Run generate_excel.py first.")
        return

    print(f"Uploading {len(files)} file(s) from input_folder/\n")
    for fname in files:
        path = os.path.join(INPUT_FOLDER, fname)
        pinfo = next((p for p in FILE_PATTERNS if re.search(p["pattern"], fname.lower())), None)
        if not pinfo:
            print(f"  Skipped '{fname}' — no pattern match")
            continue
        wb = openpyxl.load_workbook(path)
        sheets = [s for s in wb.sheetnames if re.search(pinfo["sheet_regex"], s, re.IGNORECASE)]
        if not sheets:
            print(f"  Skipped '{fname}' — sheet '{pinfo['sheet_regex']}' not found")
            continue
        data = read_sheet(path, sheets[0], pinfo["header_row"])
        col  = get_col(pinfo["collection"])
        if pinfo.get("sync"):
            sync_col(data, col, pinfo["unique_field"])
        else:
            insert_fresh(data, col)

    print("\nDone.")


if __name__ == "__main__":
    main()
