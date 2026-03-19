"""requirements_to_db.py — Base uploader for 'requirements'. Clears then inserts."""
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
import openpyxl
from pymongo import MongoClient
from datetime import datetime

MONGO_URI = "mongodb://localhost:27017/"; DB_NAME = "test_db"; COLLECTION = "requirements"
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "input_folder", "requirements_data.xlsx")
SHEET_NAME = "Requirements"; HEADER_ROW = 1

def sanitize_key(k): return k.replace(".", "_").replace("$", "_") if k else "col"
def read_sheet(path, sheet, hr):
    ws = openpyxl.load_workbook(path)[sheet]
    hdrs = [sanitize_key(str(c.value)) if c.value else f"col{i}" for i, c in enumerate(ws[hr])]
    hdrs.append("date")
    return [dict(zip(hdrs, list(row) + [datetime.now()])) for row in ws.iter_rows(min_row=hr+1, values_only=True)]

def main():
    data = read_sheet(EXCEL_FILE, SHEET_NAME, HEADER_ROW)
    col  = MongoClient(MONGO_URI)[DB_NAME][COLLECTION]
    col.delete_many({}); col.insert_many(data) if data else None
    print(f"[{COLLECTION}]  Inserted {len(data)} records")

if __name__ == "__main__": main()
