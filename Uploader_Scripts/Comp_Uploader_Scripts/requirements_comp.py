"""
requirements_comp.py  —  Compare-and-update uploader for 'requirements' by requirement_id.
This is the primary script for requirements since it tracks lifecycle via db_status.
  db_status='added'   → new record not previously in MongoDB
  db_status='updated' → record existed and was refreshed
  db_status='deleted' → record was in MongoDB but no longer in the Excel file

Usage:
    python Uploader_Scripts/Comp_Uploader_Scripts/requirements_comp.py
"""

import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

import openpyxl
from pymongo import MongoClient
from datetime import datetime

MONGO_URI    = "mongodb://localhost:27017/"
DB_NAME      = "test_db"
COLLECTION   = "requirements"
UNIQUE_FIELD = "requirement_id"
EXCEL_FILE   = os.path.join(os.path.dirname(__file__), "..", "..", "input_folder", "requirements_data.xlsx")
SHEET_NAME   = "Requirements"
HEADER_ROW   = 1


def sanitize_key(k):
    return k.replace(".", "_").replace("$", "_") if k else "col"


def read_sheet(path, sheet, hr):
    ws   = openpyxl.load_workbook(path)[sheet]
    hdrs = [sanitize_key(str(c.value)) if c.value else f"col{i}" for i, c in enumerate(ws[hr])]
    hdrs.append("date")
    return [dict(zip(hdrs, list(row) + [datetime.now()])) for row in ws.iter_rows(min_row=hr+1, values_only=True)]


def sync(data, col, uid):
    existing = {d[uid]: d for d in col.find({}, {uid: 1, "_id": 0}) if d.get(uid)}
    seen, added, updated = set(), 0, 0
    for rec in data:
        key = rec.get(uid)
        if not key:
            continue
        if key in existing:
            col.update_one({uid: key}, {"$set": {**rec, "db_status": "updated"}}); updated += 1
        else:
            col.insert_one({**rec, "db_status": "added"}); added += 1
        seen.add(key)
    deleted = col.update_many(
        {uid: {"$nin": list(seen)}}, {"$set": {"db_status": "deleted"}}
    ).modified_count
    print(f"[{col.name}]  {added} added | {updated} updated | {deleted} deleted")


def main():
    data = read_sheet(EXCEL_FILE, SHEET_NAME, HEADER_ROW)
    col  = MongoClient(MONGO_URI)[DB_NAME][COLLECTION]
    sync(data, col, UNIQUE_FIELD)


if __name__ == "__main__":
    main()
