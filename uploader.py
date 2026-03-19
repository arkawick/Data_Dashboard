"""
uploader.py  —  GraphRAG Dashboard Excel Uploader
Tkinter GUI: browse a file or folder, press Start, and all matching
Excel files are uploaded to MongoDB (test_db).

Collection routing is driven by config.py FILE_PATTERNS.
  sync=False  →  collection is cleared then re-inserted (fast, idempotent)
  sync=True   →  records are compared by unique_field; missing ones are
                 marked db_status="deleted", new ones "added", changed ones "updated"
                 (currently used for the 'requirements' collection)
"""

import os
import re
import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext

import openpyxl
from pymongo import MongoClient
from datetime import datetime

from config import FILE_PATTERNS

# ── MongoDB ───────────────────────────────────────────────────────────────────
MONGO_URI = "mongodb://localhost:27017/"
DB_NAME   = "test_db"

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_FOLDER = os.path.join(SCRIPT_DIR, "input_folder")
os.makedirs(INPUT_FOLDER, exist_ok=True)


# ── GUI setup ─────────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("GraphRAG Dashboard — Excel Uploader")
root.geometry("720x440")

status_label = tk.Label(root, text="Select a file / folder or leave blank for input_folder/",
                        font=("Arial", 10))
status_label.pack(pady=5)

frame = tk.Frame(root)
frame.pack(pady=5)

file_path_entry = tk.Entry(frame, width=55)
file_path_entry.pack(side=tk.LEFT, padx=5)


def browse_path():
    path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not path:
        path = filedialog.askdirectory()
    if path:
        file_path_entry.delete(0, tk.END)
        file_path_entry.insert(0, path)


browse_button = tk.Button(frame, text="Browse", command=browse_path)
browse_button.pack(side=tk.LEFT)

progress = ttk.Progressbar(root, orient="horizontal", length=440, mode="determinate")
progress.pack(pady=5)

log_box = scrolledtext.ScrolledText(root, height=12, width=88, state="disabled")
log_box.pack(pady=5)


def log(msg):
    log_box.config(state="normal")
    log_box.insert(tk.END, msg + "\n")
    log_box.config(state="disabled")
    log_box.yview(tk.END)


# ── Core helpers ──────────────────────────────────────────────────────────────
def get_collection(collection_name):
    client = MongoClient(MONGO_URI)
    return client[DB_NAME][collection_name]


def sanitize_key(key):
    """Replace MongoDB-invalid characters in field names."""
    return key.replace(".", "_").replace("$", "_") if key else "unknown"


def sanitize_headers(headers):
    """Sanitize and deduplicate column headers."""
    seen, result = {}, []
    for h in headers:
        s = sanitize_key(str(h)) if h is not None else "col"
        if s in seen:
            seen[s] += 1
            s = f"{s}_{seen[s]}"
        else:
            seen[s] = 0
        result.append(s)
    return result


def read_sheet(filepath, sheet_name, header_row=1):
    """Read an Excel sheet and return (headers, list-of-dicts) with a 'date' timestamp."""
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb[sheet_name]
    hdrs = sanitize_headers([c.value for c in ws[header_row]])
    hdrs.append("date")
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)
        vals.append(datetime.now())
        rows.append(dict(zip(hdrs, vals)))
    return hdrs, rows


def insert_fresh(data, collection):
    """Drop existing documents and insert all new ones."""
    collection.delete_many({})
    if data:
        collection.insert_many(data)
    return f"Inserted {len(data)} records into '{collection.name}'"


def sync_collection(data, collection, unique_field):
    """
    Compare incoming data with existing records:
      - new record       → insert with db_status='added'
      - existing record  → update with db_status='updated'
      - missing record   → mark with db_status='deleted'
    """
    existing = {doc[unique_field]: doc for doc in collection.find({}, {unique_field: 1, "_id": 0})
                if doc.get(unique_field)}
    seen_ids = set()
    added = updated = 0

    for rec in data:
        uid = rec.get(unique_field)
        if not uid:
            continue
        if uid in existing:
            collection.update_one({unique_field: uid}, {"$set": {**rec, "db_status": "updated"}})
            updated += 1
        else:
            collection.insert_one({**rec, "db_status": "added"})
            added += 1
        seen_ids.add(uid)

    deleted = collection.update_many(
        {unique_field: {"$nin": list(seen_ids)}},
        {"$set": {"db_status": "deleted"}}
    ).modified_count

    return (f"Synced '{collection.name}': "
            f"{added} added, {updated} updated, {deleted} marked deleted")


# ── File processor ────────────────────────────────────────────────────────────
def process_file(filepath):
    filename = os.path.basename(filepath).lower()

    pattern_info = next(
        (p for p in FILE_PATTERNS if re.search(p["pattern"], filename)), None
    )
    if not pattern_info:
        return f"Skipped '{os.path.basename(filepath)}' — no matching pattern in config.py"

    collection_name = pattern_info["collection"]
    sheet_regex     = pattern_info["sheet_regex"]
    header_row      = pattern_info["header_row"]
    unique_field    = pattern_info.get("unique_field")
    do_sync         = pattern_info.get("sync", False)

    try:
        wb = openpyxl.load_workbook(filepath)
        matching = [s for s in wb.sheetnames if re.search(sheet_regex, s, re.IGNORECASE)]
        if not matching:
            return f"Error: no sheet matching '{sheet_regex}' in '{os.path.basename(filepath)}'"
        _, data = read_sheet(filepath, matching[0], header_row)
    except Exception as exc:
        return f"Read error in '{os.path.basename(filepath)}': {exc}"

    try:
        col = get_collection(collection_name)
        if do_sync and unique_field:
            return sync_collection(data, col, unique_field)
        return insert_fresh(data, col)
    except Exception as exc:
        return f"MongoDB error for '{collection_name}': {exc}"


def find_excel_files(path):
    if os.path.isfile(path) and path.endswith(".xlsx"):
        return [path]
    if os.path.isdir(path):
        return sorted(
            os.path.join(path, f) for f in os.listdir(path) if f.endswith(".xlsx")
        )
    # fallback: input_folder/
    return sorted(
        os.path.join(INPUT_FOLDER, f)
        for f in os.listdir(INPUT_FOLDER) if f.endswith(".xlsx")
    )


# ── Start button ──────────────────────────────────────────────────────────────
def start_processing():
    path = file_path_entry.get().strip()
    excel_files = find_excel_files(path)

    if not excel_files:
        status_label.config(text="No Excel files found!", fg="red")
        log("No Excel files found. Run generate_excel.py first.")
        return

    status_label.config(text="Processing...", fg="blue")
    log(f"Found {len(excel_files)} file(s):")
    progress["maximum"] = len(excel_files)

    for i, fpath in enumerate(excel_files, 1):
        log(f"  [{i}/{len(excel_files)}] {os.path.basename(fpath)}")
        result = process_file(fpath)
        log(f"    => {result}")
        status_label.config(text=f"Processing: {os.path.basename(fpath)}")
        progress["value"] = i
        root.update_idletasks()

    status_label.config(text="Done!", fg="green")
    log("All files processed.")


start_button = tk.Button(root, text="Start Upload", font=("Arial", 12),
                         bg="#007BFF", fg="white", command=start_processing)
start_button.pack(pady=10)

root.mainloop()
