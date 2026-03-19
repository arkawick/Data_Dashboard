"""
generate_excel.py
Generates randomized GraphRAG-ready Excel files into input_folder/.
Run this to create fresh Excel files, then use uploader.py (or Comp scripts)
to push them into MongoDB.

Files created:
  input_folder/employees_data.xlsx     sheet: Employees
  input_folder/projects_data.xlsx      sheet: Projects
  input_folder/test_cases_data.xlsx    sheet: TestCases
  input_folder/bugs_data.xlsx          sheet: Bugs
  input_folder/requirements_data.xlsx  sheet: Requirements

List fields (skills, tech_stack) are stored as comma-separated strings in Excel.
The 'date' upload-timestamp is NOT included — uploader.py appends it automatically.
"""

import os
import random
import string
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_FOLDER = os.path.join(SCRIPT_DIR, "input_folder")
os.makedirs(INPUT_FOLDER, exist_ok=True)

# ── Lookup tables (same as generate_data.py) ──────────────────────────────────
DOMAINS       = ["Automotive", "Finance", "Healthcare", "Retail", "Industrial", "Aerospace", "Telecom"]
TEAMS         = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon"]
DEPARTMENTS   = ["Engineering", "QA", "DevOps", "Product", "Security"]
ROLES         = ["Engineer", "Senior Engineer", "Tech Lead", "QA Engineer", "DevOps Engineer", "Architect", "Product Manager"]
SENIORITIES   = ["Junior", "Mid", "Senior", "Principal", "Staff"]
LOCATIONS     = ["New York", "London", "Berlin", "Tokyo", "Sydney", "Bangalore", "Toronto"]
SKILLS_POOL   = ["Python", "Java", "C++", "JavaScript", "MongoDB", "Docker", "Kubernetes",
                 "CI/CD", "Selenium", "pytest", "JUnit", "REST API", "GraphQL", "Spark", "Redis"]
TECH_STACKS   = ["Python/FastAPI", "Java/Spring", "Node.js/Express", "Go/gRPC",
                 "C++/Qt", "React/TypeScript", "Vue.js/Django"]
PROJECT_NAMES = [
    "Platform Modernization", "API Gateway", "Data Pipeline", "Auth Service Overhaul",
    "Mobile SDK", "Analytics Engine", "Security Audit", "Cloud Migration",
    "Real-time Monitoring", "ML Inference Service"
]
PROJECT_STATUSES   = ["Active", "Completed", "On Hold", "Planning", "Cancelled"]
PROJECT_PRIORITIES = ["Critical", "High", "Medium", "Low"]
TC_STATUSES        = ["Passed", "Failed", "Skipped", "Pending", "Blocked"]
TC_TYPES           = ["Functional", "Regression", "Smoke", "Performance", "Security", "Integration"]
AUTOMATION_STATUSES = ["Automated", "Manual", "In Progress"]
PARENT_FOLDERS     = ["Unit Tests", "Integration Tests", "E2E Tests", "Performance Tests", "Security Tests"]
TC_VERBS           = ["Validate", "Verify", "Test", "Check", "Assert", "Ensure"]
TC_SUBJECTS        = [
    "login_flow", "data_ingestion", "api_response", "ui_rendering", "auth_token",
    "database_query", "cache_invalidation", "error_handling", "retry_logic",
    "rate_limiting", "session_management", "file_upload", "export_function",
    "search_feature", "pagination_logic"
]
BUG_SEVERITIES = ["Critical", "Major", "Minor", "Trivial"]
BUG_PRIORITIES = ["P1", "P2", "P3", "P4"]
BUG_STATUSES   = ["Open", "In Progress", "Resolved", "Closed", "Reopened"]
BUG_TYPES      = ["Functional", "Performance", "Security", "UI", "Data", "Concurrency"]
BUG_PREFIXES   = [
    "Crash in", "Memory leak in", "Incorrect output from", "Timeout in",
    "Race condition in", "Null pointer in", "Missing validation in",
    "Broken redirect in", "Data corruption in", "UI glitch in",
    "Performance degradation in", "Auth bypass in"
]
REQ_CATEGORIES = ["Functional", "Non-Functional", "Security", "Performance", "Compliance", "Usability"]
REQ_PRIORITIES = ["Must Have", "Should Have", "Could Have", "Won't Have"]
REQ_STATUSES   = ["Approved", "Draft", "Under Review", "Rejected", "Implemented"]
REQ_SUBJECTS   = [
    "user authentication", "data encryption", "API response time", "error logging",
    "access control", "audit trail", "data backup", "session timeout",
    "input validation", "rate limiting", "multi-factor authentication",
    "GDPR compliance", "load balancing", "monitoring dashboard", "automated alerts"
]
FIRST_NAMES = ["Alice", "Bob", "Charlie", "Diana", "Eve", "Frank", "Grace", "Hank",
               "Iris", "Jack", "Karen", "Leo", "Mia", "Noah", "Olivia", "Paul",
               "Quinn", "Rachel", "Sam", "Tina", "Uma", "Victor", "Wendy", "Xander", "Yara"]
LAST_NAMES  = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
               "Davis", "Wilson", "Taylor", "Anderson", "Thomas", "Jackson", "White",
               "Harris", "Martin", "Thompson", "Moore", "Young", "Allen"]


# ── Helpers ───────────────────────────────────────────────────────────────────
def rand_date(start_days_ago=365, end_days_ago=0):
    delta = random.randint(end_days_ago, start_days_ago)
    return (datetime.now() - timedelta(days=delta)).strftime("%Y-%m-%d")

def rand_git_hash():
    return ''.join(random.choices(string.hexdigits[:16], k=10))

def fmt_id(prefix, n, width=3):
    return f"{prefix}-{str(n).zfill(width)}"

def style_header_row(ws, n_cols):
    """Apply bold + blue fill to the first row."""
    fill = PatternFill("solid", fgColor="2196F3")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


# ── Data generators ───────────────────────────────────────────────────────────
def gen_employees(n=25):
    records, used = [], set()
    for i in range(1, n + 1):
        while True:
            fn, ln = random.choice(FIRST_NAMES), random.choice(LAST_NAMES)
            name = f"{fn} {ln}"
            if name not in used:
                used.add(name)
                break
        records.append({
            "employee_id": fmt_id("EMP", i),
            "name":        name,
            "email":       f"{fn.lower()}.{ln.lower()}@techcorp.io",
            "department":  random.choice(DEPARTMENTS),
            "role":        random.choice(ROLES),
            "team":        random.choice(TEAMS),
            "location":    random.choice(LOCATIONS),
            "seniority":   random.choice(SENIORITIES),
            "skills":      ", ".join(random.sample(SKILLS_POOL, k=random.randint(2, 5))),
            "db_status":   random.choices(["active", "inactive"], weights=[85, 15])[0],
        })
    return records


def gen_projects(n=10, employees=None):
    records = []
    for i in range(1, n + 1):
        lead  = random.choice(employees) if employees else {}
        start = rand_date(400, 100)
        records.append({
            "project_id":       fmt_id("PROJ", i),
            "project_name":     PROJECT_NAMES[i - 1],
            "domain":           random.choice(DOMAINS),
            "status":           random.choice(PROJECT_STATUSES),
            "priority":         random.choice(PROJECT_PRIORITIES),
            "tech_stack":       ", ".join(random.sample(TECH_STACKS, k=random.randint(1, 3))),
            "lead_employee_id": lead.get("employee_id", ""),
            "lead_name":        lead.get("name", ""),
            "team":             lead.get("team", random.choice(TEAMS)),
            "start_date":       start,
            "end_date":         rand_date(10, 0),
            "git_hash":         rand_git_hash(),
            "db_status":        random.choices(["active", "inactive"], weights=[85, 15])[0],
        })
    return records


def gen_test_cases(n=100, projects=None, employees=None):
    records = []
    for i in range(1, n + 1):
        proj     = random.choice(projects)  if projects  else {}
        assignee = random.choice(employees) if employees else {}
        name     = f"{random.choice(TC_VERBS)}_{random.choice(TC_SUBJECTS)}_{fmt_id('', i, 3).strip('-')}"
        records.append({
            "test_case_id":            fmt_id("TC", i),
            "test_case_name":          name,
            "domain":                  proj.get("domain", random.choice(DOMAINS)),
            "project_id":              proj.get("project_id", ""),
            "project_name":            proj.get("project_name", ""),
            "assigned_to_employee_id": assignee.get("employee_id", ""),
            "assigned_to_name":        assignee.get("name", ""),
            "parent_folder":           random.choice(PARENT_FOLDERS),
            "path_folder":             f"/{random.choice(PARENT_FOLDERS).replace(' ', '_')}/{random.choice(TC_SUBJECTS)}",
            "status":                  random.choice(TC_STATUSES),
            "test_type":               random.choice(TC_TYPES),
            "automation_status":       random.choice(AUTOMATION_STATUSES),
            "team":                    proj.get("team", random.choice(TEAMS)),
            "git_hash":                proj.get("git_hash", rand_git_hash()),
            "db_status":               random.choices(["active", "inactive"], weights=[90, 10])[0],
        })
    return records


def gen_bugs(n=150, projects=None, test_cases=None, employees=None):
    records = []
    for i in range(1, n + 1):
        proj     = random.choice(projects)   if projects   else {}
        tc       = random.choice(test_cases) if test_cases else {}
        reporter = random.choice(employees)  if employees  else {}
        assignee = random.choice(employees)  if employees  else {}
        component = random.choice(TC_SUBJECTS).replace("_", " ")
        records.append({
            "bug_id":               fmt_id("BUG", i),
            "title":                f"{random.choice(BUG_PREFIXES)} {component}",
            "description":          f"Found during {tc.get('test_case_name', 'manual testing')}",
            "severity":             random.choice(BUG_SEVERITIES),
            "priority":             random.choice(BUG_PRIORITIES),
            "status":               random.choice(BUG_STATUSES),
            "bug_type":             random.choice(BUG_TYPES),
            "project_id":           proj.get("project_id", ""),
            "project_name":         proj.get("project_name", ""),
            "test_case_id":         tc.get("test_case_id", ""),
            "test_case_name":       tc.get("test_case_name", ""),
            "reporter_employee_id": reporter.get("employee_id", ""),
            "reporter_name":        reporter.get("name", ""),
            "assignee_employee_id": assignee.get("employee_id", ""),
            "assignee_name":        assignee.get("name", ""),
            "domain":               proj.get("domain", random.choice(DOMAINS)),
            "team":                 proj.get("team", random.choice(TEAMS)),
            "git_hash":             tc.get("git_hash", rand_git_hash()),
            "db_status":            random.choices(["active", "inactive"], weights=[80, 20])[0],
        })
    return records


def gen_requirements(n=80, projects=None, test_cases=None, employees=None):
    records = []
    for i in range(1, n + 1):
        proj  = random.choice(projects)   if projects   else {}
        tc    = random.choice(test_cases) if test_cases else {}
        owner = random.choice(employees)  if employees  else {}
        subj  = random.choice(REQ_SUBJECTS)
        records.append({
            "requirement_id":              fmt_id("REQ", i),
            "requirement_name":            f"The system shall support {subj}",
            "description":                 f"Requirement for {proj.get('project_name', 'the system')} to implement {subj}",
            "domain":                      proj.get("domain", random.choice(DOMAINS)),
            "category":                    random.choice(REQ_CATEGORIES),
            "priority":                    random.choice(REQ_PRIORITIES),
            "status":                      random.choice(REQ_STATUSES),
            "project_id":                  proj.get("project_id", ""),
            "project_name":                proj.get("project_name", ""),
            "verification_responsibility": owner.get("name", ""),
            "verifier_employee_id":        owner.get("employee_id", ""),
            "covered_by_test_case_id":     tc.get("test_case_id", ""),
            "covered_by_test_case_name":   tc.get("test_case_name", ""),
            "team":                        proj.get("team", random.choice(TEAMS)),
            # db_status intentionally omitted — set by sync uploader
        })
    return records


# ── Excel writer ──────────────────────────────────────────────────────────────
def write_excel(filename, sheet_name, records):
    if not records:
        print(f"  Skipped {filename} (no data)")
        return
    path = os.path.join(INPUT_FOLDER, filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(records[0].keys())
    ws.append(headers)

    for rec in records:
        ws.append([rec.get(h, "") for h in headers])

    style_header_row(ws, len(headers))
    auto_width(ws)
    wb.save(path)
    print(f"  {filename:<35} {len(records)} rows  ->  sheet: '{sheet_name}'")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Generating Excel files into: {INPUT_FOLDER}\n")

    employees   = gen_employees(25)
    projects    = gen_projects(10, employees)
    test_cases  = gen_test_cases(100, projects, employees)
    bugs        = gen_bugs(150, projects, test_cases, employees)
    requirements = gen_requirements(80, projects, test_cases, employees)

    write_excel("employees_data.xlsx",    "Employees",    employees)
    write_excel("projects_data.xlsx",     "Projects",     projects)
    write_excel("test_cases_data.xlsx",   "TestCases",    test_cases)
    write_excel("bugs_data.xlsx",         "Bugs",         bugs)
    write_excel("requirements_data.xlsx", "Requirements", requirements)

    print("\nDone. Upload to MongoDB with:")
    print("  python uploader.py              (GUI — picks up input_folder/ automatically)")
    print("  python Uploader_Scripts/Base_Uploader_Scripts/upload_all.py  (CLI batch)")


if __name__ == "__main__":
    main()
